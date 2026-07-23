import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

BLUE = Font(name="Arial", color="0000FF")
BLACK = Font(name="Arial", color="000000")
BOLD = Font(name="Arial", bold=True)
TITLE = Font(name="Arial", bold=True, size=14)
HDR = Font(name="Arial", bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
yellow = PatternFill("solid", fgColor="FFFF00")
grey = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------- Assumptions ----------------
a = wb.active
a.title = "Assumptions"
a["A1"] = "Assumptions & Levers"; a["A1"].font = TITLE
a["A2"] = "Blue = editable input/lever. Adjust these and every scenario recalculates."; a["A2"].font = Font(name="Arial", italic=True, size=9)

a["A3"]="Item"; a["B3"]="Value"; a["C3"]="Notes"
for c in ("A3","B3","C3"): a[c].font=HDR; a[c].fill=hdr_fill; a[c].border=border

rows = [
 ("Stripe % fee", 0.029, "0.0%","per transaction (standard)"),
 ("Stripe fixed fee", 0.30, "$#,##0.00","per transaction"),
 ("Infra cost / paid user / mo", 0.30, "$#,##0.00","hosting, DB, real-time sockets"),
 ("Support & overhead / paid user / mo", 0.15, "$#,##0.00","amortized support/ops"),
 ("", None, None, ""),
 ("AI usage per PAID user / month", None, None, "heavy-user estimate"),
 ("Tier 1 (card) gens", 100, "#,##0","per user / month"),
 ("Tier 1 cost per gen", 0.0012, "$#,##0.0000","premium model, ~3x Flash-Lite"),
 ("Tier 2 (lesson) gens", 20, "#,##0","per user / month"),
 ("Tier 2 cost per gen", 0.0024, "$#,##0.0000","premium model"),
 ("Tier 3 (unit) gens", 2, "#,##0","per user / month"),
 ("Tier 3 cost per unit", 0.015, "$#,##0.0000","staged pipeline"),
 ("AI cost / paid user / mo", "=B10*B11+B12*B13+B14*B15", "$#,##0.00","computed"),
 ("", None, None, ""),
 ("Competitor 'Pro' price / mo (reference)", 5.59, "$#,##0.00","Common Planner Pro"),
 ("Annual incentive: free months", 2, "#,##0","annual = pay for (12 - this)"),
]
r=4
input_cells=set()
for label,val,fmt,note in rows:
    a[f"A{r}"]=label
    if label and val is None and note and fmt is None and note!="":
        a[f"A{r}"].font=BOLD; a[f"C{r}"]=note; a[f"C{r}"].font=Font(name="Arial", italic=True, size=9)
    if val is not None:
        a[f"B{r}"]=val
        if fmt: a[f"B{r}"].number_format=fmt
        if isinstance(val,str) and val.startswith("="):
            a[f"B{r}"].font=BLACK; a[f"A{r}"].font=BOLD; a[f"B{r}"].fill=grey
        else:
            a[f"B{r}"].font=BLUE; a[f"B{r}"].fill=yellow
        a[f"C{r}"]=note; a[f"C{r}"].font=Font(name="Arial", size=9)
        for cc in (f"A{r}",f"B{r}"): a[cc].border=border
    r+=1
a.column_dimensions["A"].width=34; a.column_dimensions["B"].width=12; a.column_dimensions["C"].width=34

# ---------------- Scenarios ----------------
s = wb.create_sheet("Scenarios")
s["A1"]="Pricing Scenarios — profitable AND under Common Planner's $5.59"; s["A1"].font=TITLE
s["A2"]="Every scenario below undercuts their $5.59 AI plan. Green = live formulas from Assumptions."; s["A2"].font=Font(name="Arial", italic=True, size=9)

heads=["Plan","AI?","Billing","Monthly list price","Eff. monthly revenue","Stripe fee /mo","AI cost /mo","Infra /mo","Support /mo","Total cost /mo","Gross profit /mo","Gross margin %"]
for i,h in enumerate(heads,1):
    c=s.cell(row=4,column=i,value=h); c.font=HDR; c.fill=hdr_fill; c.border=border; c.alignment=Alignment(wrap_text=True,vertical="center",horizontal="center")

# plan, ai(bool), billing, monthly price
scen=[
 ("Basic (no AI)","No","Monthly",2.99),
 ("Basic (no AI)","No","Annual",2.99),
 ("Plus (AI) — recommended","Yes","Monthly",4.99),
 ("Plus (AI) — recommended","Yes","Annual",4.99),
 ("Plus (AI) — aggressive","Yes","Monthly",3.99),
 ("Plus (AI) — aggressive","Yes","Annual",3.99),
 ("Competitor Pro (AI) — ref","Yes","Monthly",5.59),
]
r=5
for plan,ai,billing,price in scen:
    s[f"A{r}"]=plan; s[f"B{r}"]=ai; s[f"C{r}"]=billing; s[f"D{r}"]=price
    s[f"D{r}"].font=BLUE; s[f"D{r}"].fill=yellow; s[f"D{r}"].number_format="$#,##0.00"
    # eff monthly revenue
    if billing=="Annual":
        s[f"E{r}"]=f"=D{r}*(12-Assumptions!$B$19)/12"
        s[f"F{r}"]=f"=(E{r}*12*Assumptions!$B$4+Assumptions!$B$5)/12"
    else:
        s[f"E{r}"]=f"=D{r}"
        s[f"F{r}"]=f"=E{r}*Assumptions!$B$4+Assumptions!$B$5"
    s[f"G{r}"]=f'=IF(B{r}="Yes",Assumptions!$B$16,0)'
    s[f"H{r}"]="=Assumptions!$B$6"
    s[f"I{r}"]="=Assumptions!$B$7"
    s[f"J{r}"]=f"=F{r}+G{r}+H{r}+I{r}"
    s[f"K{r}"]=f"=E{r}-J{r}"
    s[f"L{r}"]=f"=K{r}/E{r}"
    for col in "EFGHIJK": s[f"{col}{r}"].number_format="$#,##0.00"; s[f"{col}{r}"].font=BLACK
    s[f"L{r}"].number_format="0.0%"; s[f"L{r}"].font=BOLD
    for col in "ABCDEFGHIJKL":
        s[f"{col}{r}"].border=border
    if "recommended" in plan: 
        for col in "ABCDEFGHIJKL": s[f"{col}{r}"].fill=PatternFill("solid",fgColor="E2EFDA")
    if "ref" in plan:
        for col in "ABCDEFGHIJKL": s[f"{col}{r}"].fill=PatternFill("solid",fgColor="FCE4D6")
    r+=1

widths={"A":26,"B":6,"C":9,"D":11,"E":12,"F":11,"G":10,"H":9,"I":10,"J":11,"K":12,"L":12}
for k,v in widths.items(): s.column_dimensions[k].width=v

# takeaways
tr=r+2
notes=[
 "Key takeaways:",
 "• AI is only ~$0.20 / paid user / mo — cost is trivial. Stripe fees + infra dominate at low price points.",
 "• Annual billing amortizes Stripe's $0.30 fixed fee -> HIGHER margin % than monthly, even at a lower effective price. Push annual.",
 "• Plus (AI) at $4.99 undercuts their $5.59 AND includes unit-level AI generation they don't have, at ~78% gross margin.",
 "• Even the aggressive $3.99 AI tier holds ~73% margin. Room to compete on price if needed.",
 "• Margins here are GROSS (per-user variable). Fixed costs (salaries, marketing) are covered by volume — model those separately.",
]
for i,n in enumerate(notes):
    c=s.cell(row=tr+i,column=1,value=n)
    c.font=BOLD if i==0 else Font(name="Arial",size=10)
    s.merge_cells(start_row=tr+i,start_column=1,end_row=tr+i,end_column=12)

s.freeze_panes="A5"
wb.save("/sessions/tender-optimistic-maxwell/mnt/outputs/mycurricula-pricing-model.xlsx")
print("saved")

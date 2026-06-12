#!/usr/bin/env python3
# =============================================================================
# gen-grade5-daily-sql.py — Stage-2 generator for the Grade 5 per-day rebuild.
#
# Turns scripts/_grade5-day-extract.json (471 day-level records extracted from
# the teacher workbook by scripts/extract-grade5-days.py) into an idempotent
# SQL load (scripts/_grade5-daily-load.sql) that REPLACES the 185 all-on-Sunday
# week-summary lessons from the prior load with real daily lessons.
#
# Teacher-approved mapping (binding — see the Stage-2 brief):
#   1. Literacy day cells SPLIT by inline line-anchored headings:
#        "Guided Reading…"  -> reading      "Writing:" / "Writing DAY n" -> writing
#        "Grammar:"         -> grammar      "Spelling…"                  -> spelling
#        "Phonics:"         -> ufli
#      Content before the first heading joins the first detected section.
#      A cell with no recognizable headings stays whole as one reading lesson.
#   2. SEL rows (Morning Meeting / SE Learning / Closing Circle — summarized in
#      the extract's `unmapped` list but with day cells only in the workbook)
#      load as ONE combined `sel` lesson per day. The workbook itself is
#      re-walked here for those three row families only (path comes from the
#      extract's meta.source_xlsx) because the Stage-1 JSON intentionally did
#      not materialize per-day records for unmapped rows.
#   3. math / explorers records load as-is.
#   4. Weeks 35-37 (Unit 7 Spiral Review, no workbook sheets): the 15 existing
#      week-summary lessons are NOT deleted; their day_of_week is spread
#      one-subject-per-day (math->sun, reading->mon, writing->tue,
#      grammar->wed, explorers->thu), guarded by day_of_week='sun'.
#   5. The other 170 week-summary lessons (weeks 1-34) are SOFT-deleted by an
#      explicit id list read from scripts/_grade5-real-data.sql.
#   6. day_span / merged_span records land on the FIRST day of the span with
#      "(also covers <days>)" appended to notes.
#
# Conventions mirror scripts/gen-grade5-real-data.py (canonical reference):
# slug_uuid/uuid_v5 id bridge, fixed SCHOOL/GRADE/YEAR/SUBJECT_ID uuids, the
# parse_standards approach, q()/jq() escaping, and the exact
# master_core_lesson_events column list (grade_level_id is intentionally NOT
# set — a BEFORE trigger derives it from unit_id since the 2026-06-04
# planner_scale_hardening migration; the prior SQL does not set it either).
#
# Deterministic, offline, stdlib + openpyxl only.
# =============================================================================

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

SCRIPTS_DIR = Path(__file__).resolve().parent
EXTRACT_PATH = SCRIPTS_DIR / "_grade5-day-extract.json"
PRIOR_SQL_PATH = SCRIPTS_DIR / "_grade5-real-data.sql"
OUT_PATH = SCRIPTS_DIR / "_grade5-daily-load.sql"

# ── id bridge (port of lib/planner/id-bridge.ts, identical to the prior gen) ─
import hashlib

NS = re.sub(r"[^0-9a-f]", "0", "6f1d2c84-3b6a-4e2f-9a77-planner000000").replace("-", "")
NS_BYTES = bytes.fromhex(NS[: len(NS) // 2 * 2])  # match Node odd-nibble drop
KIND_SALT = {"lesson": "lesson:", "unit": "unit:", "subject": "subject:",
             "standard": "standard:", "framework": "framework:"}


def uuid_v5(name):
    h = hashlib.sha1(NS_BYTES + name.encode("utf-8")).digest()
    b = bytearray(h[:16])
    b[6] = (b[6] & 0x0F) | 0x50
    b[8] = (b[8] & 0x3F) | 0x80
    x = b.hex()
    return f"{x[0:8]}-{x[8:12]}-{x[12:16]}-{x[16:20]}-{x[20:]}"


def slug_uuid(kind, slug):
    return uuid_v5(f"{KIND_SALT.get(kind, kind + ':')}{slug}")


# ── fixed ids from supabase/seed.sql ─────────────────────────────────────────
SCHOOL = "00000000-0000-0000-0000-0000000000a1"
GRADE = "00000000-0000-0000-0000-0000000000b5"
YEAR_2627 = "00000000-0000-0000-0000-0000000000c2"
SUBJECT_ID = {
    "math":      "00000000-0000-0000-0000-0000000005d1",
    "reading":   "00000000-0000-0000-0000-0000000005d2",
    "writing":   "00000000-0000-0000-0000-0000000005d3",
    "grammar":   "00000000-0000-0000-0000-0000000005d4",
    "spelling":  "00000000-0000-0000-0000-0000000005d5",
    "ufli":      "00000000-0000-0000-0000-0000000005d6",
    "explorers": "00000000-0000-0000-0000-0000000005d7",
    "sel":       "00000000-0000-0000-0000-0000000005d8",
}
SUBJECT_ID_REV = {v: k for k, v in SUBJECT_ID.items()}
# Stable per-(week, day) card order for the new daily lessons (Stage-2 brief).
DISPLAY_ORDER = {"math": 0, "reading": 1, "writing": 2, "grammar": 3,
                 "spelling": 4, "ufli": 5, "explorers": 6, "sel": 7}
# Subjects whose rows must exist before the lesson inserts (seed.sql values,
# mirrored verbatim; all three already exist in prod via seed-cloud.sql, so
# these are idempotent insurance only).
NEW_SUBJECTS = [
    ("spelling", "Spelling", 4),
    ("ufli", "UFLI", 5),
    ("sel", "SEL", 7),
]

DAY_TOKENS = ["sun", "mon", "tue", "wed", "thu", "fri", "sat"]
DAY_FULL = {"sun": "Sunday", "mon": "Monday", "tue": "Tuesday",
            "wed": "Wednesday", "thu": "Thursday", "fri": "Friday", "sat": "Saturday"}
# Thematic unit bands shared by every subject (mirrors the prior load's units).
BANDS = [(1, 1, 5), (2, 6, 12), (3, 13, 18), (4, 19, 22), (5, 23, 26), (6, 27, 34)]


def band_of_week(gw):
    for b, a, z in BANDS:
        if a <= gw <= z:
            return b
    raise ValueError(f"week {gw} outside bands 1-34")


# ── SQL escaping (port of the prior generator) ───────────────────────────────
def q(s):
    return "'" + str(s).replace("'", "''") + "'"


def jq(o):
    return "'" + json.dumps(o, ensure_ascii=False).replace("'", "''") + "'::jsonb"


# ── standards parsing (port of the prior generator, piece-for-piece) ─────────
CODE_RE = re.compile(r"^(?:EE\.?)?[A-Za-z]{1,5}\.?\d[A-Za-z0-9.]*$")


def normalize_code(c):
    c = c.strip()
    if c.upper().startswith("CCSS"):
        c = re.sub(r"^CCSS[.:]?\s*", "", c, flags=re.I)
    elif c.upper().startswith("CC") and len(c) > 2 and c[2].isalpha():
        c = c[2:]
    return c.strip(" .")


def parse_standards_pieces(std_text):
    """Prior parse_standards, but ALSO returns the leftover pieces that did not
    yield a valid-shaped code, so nothing is ever silently dropped (Stage-2
    brief: unmatched standards text goes to notes verbatim)."""
    codes, leftovers = [], []
    for piece in re.split(r"[,\n]", std_text):
        raw = piece.strip()
        if not raw:
            continue
        p = re.sub(r"^(CCSS|CC|Std)\s*:\s*", "", raw, flags=re.I).strip()
        p = normalize_code(p)
        if (p and " " not in p
                and any(c.isdigit() for c in p) and any(c.isalpha() for c in p)
                and CODE_RE.match(p)):
            codes.append(p)
        else:
            leftovers.append(raw)
    seen, out = set(), []
    for c in codes:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out, leftovers


# ── prior-load ground truth (read from the committed SQL, per the brief) ─────
def load_prior_sql():
    sql = PRIOR_SQL_PATH.read_text(encoding="utf-8")
    lesson_re = re.compile(
        r"insert into master_core_lesson_events\s*\n\s*\([^)]+\)\s*\n\s*values \("
        r"'([0-9a-f-]{36})', '([0-9a-f-]{36})', '([0-9a-f-]{36})', (\d+), '(\w+)'::weekday", re.S)
    lessons = [(i, u, s, int(w), d) for i, u, s, w, d in lesson_re.findall(sql)]
    if len(lessons) != 185:
        sys.exit(f"expected 185 prior lesson inserts, parsed {len(lessons)}")
    # cross-check every id against the slug scheme (belt and braces)
    for i, _, s, w, _ in lessons:
        slug = f"g5-2627-{SUBJECT_ID_REV[s]}-w{w}"
        if slug_uuid("lesson", slug) != i:
            sys.exit(f"prior id {i} does not match slug {slug}")
    soft_delete_ids = [i for i, _, _, w, _ in lessons if w <= 34]
    unit7 = [(i, SUBJECT_ID_REV[s], w) for i, _, s, w, _ in lessons if w >= 35]
    if len(soft_delete_ids) != 170 or len(unit7) != 15:
        sys.exit(f"expected 170 + 15 prior ids, got {len(soft_delete_ids)} + {len(unit7)}")

    std_re = re.compile(
        r"insert into standards \(id, framework_id, grade_level_id, code\)\s*\n\s*values "
        r"\('([0-9a-f-]{36})', '[0-9a-f-]{36}', '[0-9a-f-]{36}', '((?:[^']|'')*)'\)")
    standards_map = {}
    for sid, code in std_re.findall(sql):
        code = code.replace("''", "'")
        if slug_uuid("standard", code) != sid:
            sys.exit(f"prior standard id {sid} does not match code {code}")
        standards_map[code] = sid
    if len(standards_map) != 108:
        sys.exit(f"expected 108 prior standards, parsed {len(standards_map)}")

    unit_re = re.compile(
        r"insert into units \(id, grade_level_id, subject_id, school_year_id, name, "
        r"start_week, end_week\)\s*\n\s*values \('([0-9a-f-]{36})', '[0-9a-f-]{36}', "
        r"'([0-9a-f-]{36})', '[0-9a-f-]{36}', '((?:[^']|'')*)', (\d+), (\d+)\)")
    units = {}        # (subject_token, band_index) -> unit uuid
    band_names = {}   # band_index -> thematic band name (from the explorers unit)
    for uid, sid, name, a, z in unit_re.findall(sql):
        subj = SUBJECT_ID_REV[sid]
        band = next((b for b, ba, bz in BANDS if int(a) == ba and int(z) == bz), None)
        if band is None:
            continue  # the week-35-37 unit-7 rows — not a lesson target here
        units[(subj, band)] = uid
        if subj == "explorers":
            band_names[band] = name.replace("''", "'")
    if len(band_names) != 6:
        sys.exit(f"expected 6 thematic band names, got {len(band_names)}")
    return soft_delete_ids, unit7, standards_map, units, band_names


# ── literacy section splitter ────────────────────────────────────────────────
# Line-anchored heading variants, derived by scanning every literacy cell_text
# in the extract (the run report prints the matched-heading census so coverage
# is auditable). Deliberately conservative — evidence-backed exclusions:
#   * bare "Writing" lines are rotation-station labels ("RazKids, EPIC, Book
#     Bag" / "Kahoot" neighbors), never section starts;
#   * "Padlet for Spelling" is always content under a "Spelling Padlet:" line;
#   * "(15 mins) Guided Writing:" is a timed sub-segment INSIDE a Writing
#     section; "(10 mins) …Writing…" likewise;
#   * "No prep phonics:" / "Decodable text hub" are rotation/resource lines
#     inside an open section;
#   * "Guided Reading Groups will be completed." (no colon/slash) is a status
#     sentence, so guided-reading requires ':' or '/' right after the keyword;
#   * 5A/5B/5C differentiation lines ("5A-Using Ufli Words", "5B-Grade 4
#     Spelling words") live inside Spelling sections and must not split them.
HEADING_RULES = [
    # combined ELA headings always start with Guided Reading -> reading owns them
    ("guided_reading", "reading", re.compile(r"^guided\s*read(?:ing)?\s*[:/]", re.I)),
    ("writing",        "writing", re.compile(r"^writing\s*:", re.I)),
    ("writing_day_n",  "writing", re.compile(r"^writing\s+day\s*\d", re.I)),
    ("grammar",        "grammar", re.compile(r"^grammar\s*:", re.I)),
    ("spelling",       "spelling", re.compile(r"^spelling\s*:", re.I)),
    ("spelling_padlet", "spelling", re.compile(r"^spelling\s+padlet\s*:", re.I)),
    ("spelling_words", "spelling", re.compile(r"^spelling\s+words\b", re.I)),
    ("phonics",        "ufli", re.compile(r"^phonics\s*:", re.I)),
]


def match_heading(line):
    s = line.strip()
    for variant, subject, rx in HEADING_RULES:
        if rx.match(s):
            return variant, subject
    return None


def split_literacy_cell(cell_text):
    """-> (ordered [(subject, section_text, heading_line)], heading_census,
           combined_census). Preamble joins the first detected section; cells
       with no headings return a single ('reading', whole_cell, None)."""
    lines = cell_text.split("\n")
    census, combos = Counter(), Counter()
    sections = []          # [{subject, head_idx, lines}]
    preamble = []
    current = None
    for ln in lines:
        hit = match_heading(ln)
        if hit:
            variant, subject = hit
            census[variant] += 1
            if variant == "guided_reading":
                low = ln.lower()
                others = [t for t in ("grammar", "writing") if t in low]
                if others:
                    combos["guided_reading+" + "+".join(others)] += 1
            current = {"subject": subject, "heading": ln.strip(), "lines": [ln]}
            sections.append(current)
        elif current is not None:
            current["lines"].append(ln)
        else:
            preamble.append(ln)
    if not sections:
        return [("reading", cell_text, None)], census, combos
    pre = "\n".join(preamble).strip()
    out = []               # one lesson per subject; same-subject sections merge
    by_subject = {}
    for i, sec in enumerate(sections):
        text = "\n".join(sec["lines"]).strip()
        if i == 0 and pre:
            text = pre + "\n\n" + text
        if sec["subject"] in by_subject:
            slot = by_subject[sec["subject"]]
            slot[1] = slot[1] + "\n\n" + text
        else:
            slot = [sec["subject"], text, sec["heading"]]
            by_subject[sec["subject"]] = slot
            out.append(slot)
    return [tuple(s) for s in out], census, combos


# ── field shaping ────────────────────────────────────────────────────────────
BULLET_RE = re.compile(r"^[\s•◦*\-–—·]+")


def make_title(first_line, fallback="Lesson"):
    """Stage-2 brief: title = first non-empty line, leading bullets/asterisks
    stripped, trimmed to 80 chars (word-boundary cut + ellipsis)."""
    t = BULLET_RE.sub("", first_line.strip())
    t = re.sub(r"\s{2,}", " ", t).strip().rstrip(":").strip()
    if not t:
        return fallback
    if len(t) > 80:
        t = t[:79].rsplit(" ", 1)[0].rstrip(" ,;:-–—") + "…"
    return t


def first_nonempty_line(text):
    for ln in text.split("\n"):
        if ln.strip():
            return ln
    return ""


def objectives_list(text):
    """meta.objectives blob -> jsonb array of strings (one per line, leading
    bullets and OBJ:/Objective: tags stripped)."""
    out = []
    for ln in text.split("\n"):
        s = BULLET_RE.sub("", ln.strip())
        s = re.sub(r"^(OBJ|Objectives?)\b\s*[:.\-]\s*", "", s, flags=re.I).strip()
        if s:
            out.append(s)
    return out


URL_RE = re.compile(r"https?://[^\s)>\]]+")
PROVIDER = [("docs.google.com", "gdocs"), ("drive.google", "gdrive")]


def label_note_resources(note):
    """label_note URLs -> resources jsonb (prior load's shape); returns
    (resources, remaining_text)."""
    urls = URL_RE.findall(note)
    if not urls:
        return [], note.strip()
    label_src = URL_RE.sub("", note)
    label = re.sub(r"\s{2,}", " ", label_src).strip(" :-–—")
    label = (label[:77] + "…") if len(label) > 80 else (label or "Link")
    res = []
    for u in urls:
        u = u.rstrip(".,;")
        prov = next((p for frag, p in PROVIDER if frag in u), "website")
        res.append({"type": "doc" if prov in ("gdocs", "gdrive") else "link",
                    "label": label, "url": u, "provider": prov,
                    "displayMode": "hyperlink"})
    # if stripping URLs still leaves substantial guidance text, keep it in notes
    remaining = label_src.strip()
    return res, remaining if len(remaining) > 100 else ""


def span_days_note(first_tok, covered, label=None):
    """'(also covers Monday, Tuesday)' — covered excludes the placement day."""
    rest = [DAY_FULL[t] for t in covered if t != first_tok]
    if not rest:
        return ""
    days = ", ".join(rest)
    return f"({label} also covers {days})" if label else f"(also covers {days})"


def with_kind_prefix(kind_label, text):
    """'Assessment: x' — skip the prefix when the text already leads with it."""
    if re.match(rf"^\s*{re.escape(kind_label.split('/')[0])}", text, flags=re.I):
        return text
    return f"{kind_label}: {text}"


def build_notes_and_standards(meta, extra_note_parts, standards_map, stats):
    """meta dict -> (notes_text, standards_uuid_list, objectives_list)."""
    notes = list(extra_note_parts)
    std_uuids, objectives = [], []
    if meta:
        if meta.get("objectives"):
            objectives = objectives_list(meta["objectives"])
        if meta.get("standards"):
            codes, leftovers = parse_standards_pieces(meta["standards"])
            unmatched = False
            for c in codes:
                if c in standards_map:
                    if standards_map[c] not in std_uuids:
                        std_uuids.append(standards_map[c])
                    stats["std_matched"] += 1
                else:
                    unmatched = True      # valid shape, but not in the prior map
                    stats["std_unmatched_code"] += 1
            # Anything that did not resolve to a prior-load standard row keeps
            # the FULL original standards text in notes, verbatim — never
            # silently dropped, and never reflowed by the comma/newline split.
            if leftovers or unmatched:
                stats["std_text_kept_in_notes"] += 1
                notes.append(with_kind_prefix("Std", meta["standards"]))
        if meta.get("assessment"):
            notes.append(with_kind_prefix("Assessment", meta["assessment"]))
        if meta.get("differentiation"):
            notes.append(with_kind_prefix("Differentiation", meta["differentiation"]))
        if meta.get("plg_islamic_values"):
            notes.append(with_kind_prefix("PLG/Islamic Values", meta["plg_islamic_values"]))
    return "\n\n".join(p for p in notes if p), std_uuids, objectives


# ── SEL recovery from the workbook (Morning Meeting / SE Learning / Closing
#    Circle day cells — the Stage-1 JSON only summarizes these rows) ──────────
DAY_NAMES = {"sunday": "sun", "monday": "mon", "tuesday": "tue",
             "wednesday": "wed", "thursday": "thu", "friday": "fri", "saturday": "sat"}
SEL_KEYWORDS = [
    (re.compile(r"^morning meeting\b", re.I), "Morning Meeting"),
    (re.compile(r"^se learning\b", re.I), "SE Learning"),
    (re.compile(r"^closing circle\b", re.I), "Closing Circle"),
]
SEL_ORDER = ["Morning Meeting", "SE Learning", "Closing Circle"]
SUBJECT_LABEL_RES = [re.compile(r"^literacy\b", re.I), re.compile(r"^math\b", re.I),
                     re.compile(r"^explore\b", re.I), re.compile(r"^ipc\b", re.I)]
META_LABELS = {
    "learning goals objectives": "objectives", "learning goals/objectives": "objectives",
    "learning objectives": "objectives", "objective": "objectives", "objectives": "objectives",
    "assessment": "assessment", "assessments": "assessment",
    "standard": "standards", "standards": "standards",
    "differentiation": "differentiation",
    "personal learning goal/islamic values": "plg_islamic_values",
    "ipc personal learning goal/islamic values": "plg_islamic_values",
}
META_CELL_PREFIX = re.compile(
    r"^\s*(assessment|assessments|obj|objective|differentiation|std|standard|plg)\b[\s:.-]", re.I)
META_PREFIX_KIND = {"assessment": "assessment", "assessments": "assessment",
                    "obj": "objectives", "objective": "objectives",
                    "differentiation": "differentiation",
                    "std": "standards", "standard": "standards",
                    "plg": "plg_islamic_values"}


def collapse(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def clean_text(s):
    if s is None:
        return ""
    t = str(s).replace("\r\n", "\n").replace("\r", "\n")
    t = "\n".join(ln.rstrip() for ln in t.split("\n"))
    return re.sub(r"\n{3,}", "\n\n", t).strip()


def extract_sel(xlsx_path, extract_meta):
    """-> {gw: {day_token: {label: {"text": str, "covered": [tokens],
                                    "meta": {kind: text}}}}}
    Ports the Stage-1 extractor's row walk (merge handling, blank-row
    continuations, metadata sub-rows) but materializes ONLY the three SEL row
    families. Fails closed if the workbook's structure has drifted from the
    frozen Stage-1 extract (same sheet census, unit week counts, and
    global-week mapping), so SEL lessons can never come from a different
    source version than the math/literacy/explorers records."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install --user openpyxl")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    week_re = re.compile(r"^U(\d+) W(\d+)$")
    sheets = sorted((int(m.group(1)), int(m.group(2)), n)
                    for n in wb.sheetnames if (m := week_re.match(n)))
    weeks_per_unit = Counter(u for u, _, _ in sheets)
    offsets, acc = {}, 0
    for u in sorted(weeks_per_unit):
        offsets[u] = acc
        acc += weeks_per_unit[u]
    # ── drift guard: workbook must match the frozen extract's structure ──────
    if len(sheets) != extract_meta["week_sheet_count"]:
        sys.exit(f"workbook drift: {len(sheets)} week sheets vs "
                 f"{extract_meta['week_sheet_count']} in the Stage-1 extract — "
                 "re-run scripts/extract-grade5-days.py first")
    for u_str, info in extract_meta["units"].items():
        u = int(u_str)
        if weeks_per_unit.get(u) != info["weeks"] or \
                [offsets[u] + 1, offsets[u] + weeks_per_unit[u]] != info["global_weeks"]:
            sys.exit(f"workbook drift: unit {u} weeks/global-week mapping no longer "
                     "matches the Stage-1 extract — re-run scripts/extract-grade5-days.py")
    if [1, acc] != extract_meta["global_weeks_covered"]:
        sys.exit("workbook drift: global week range no longer matches the Stage-1 extract")

    out = defaultdict(lambda: defaultdict(dict))
    for u, w, sn in sheets:
        ws = wb[sn]
        gw = offsets[u] + w
        hdr = next((row[0].row for row in ws.iter_rows(min_row=1, max_row=25, max_col=1)
                    if row[0].value and collapse(row[0].value).lower() == "class"), None)
        if hdr is None:
            continue
        day_cols = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(row=hdr, column=c).value
            if v is None or not collapse(v):
                continue
            name = collapse(v).lower()
            toks = sorted((t for dn, t in DAY_NAMES.items() if dn in name),
                          key=DAY_TOKENS.index)
            if toks:
                day_cols.append((c, toks))
        amap, merges = {}, list(ws.merged_cells.ranges)
        for mr in merges:
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    amap[(r, c)] = (mr.min_row, mr.min_col)

        def cell_val(r, c):
            if (r, c) in amap:
                ar, ac = amap[(r, c)]
                mr = next(m for m in merges
                          if m.min_row == ar and m.min_col == ac
                          and m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col)
                return ws.cell(row=ar, column=ac).value, ar, ac, mr
            return ws.cell(row=r, column=c).value, r, c, None

        cur_label = None  # SEL canonical label while inside an SEL block; else None

        def emit(row, label, as_meta_kind=None):
            seen = set()
            for c, toks in day_cols:
                v, ar, ac, mr = cell_val(row, c)
                if v is None or not clean_text(v) or ar != row or (ar, ac) in seen:
                    continue
                seen.add((ar, ac))
                covered = ([t for cc, tt in day_cols for t in tt
                            if mr.min_col <= cc <= mr.max_col] if mr is not None
                           else list(toks))
                first = toks[0]
                slot = out[gw][first].setdefault(
                    label, {"text": "", "covered": list(covered), "meta": {}})
                txt = clean_text(v)
                if as_meta_kind:
                    prev = slot["meta"].get(as_meta_kind)
                    slot["meta"][as_meta_kind] = (prev + "\n\n" + txt) if prev else txt
                else:
                    slot["text"] = (slot["text"] + "\n\n" + txt) if slot["text"] else txt
                    for t in covered:
                        if t not in slot["covered"]:
                            slot["covered"].append(t)

        for r in range(hdr + 1, ws.max_row + 1):
            raw, lar, lac, lmr = cell_val(r, 1)
            if lar != r:
                raw = None  # label merged down: row continues the open block
            if raw is None or not collapse(raw):
                if cur_label is None:
                    continue
                cells = [clean_text(cell_val(r, c)[0]) for c, _ in day_cols
                         if cell_val(r, c)[1] == r and cell_val(r, c)[0] is not None
                         and clean_text(cell_val(r, c)[0])]
                if not cells:
                    continue
                pm = META_CELL_PREFIX.match(cells[0])
                if pm and all(META_CELL_PREFIX.match(x) for x in cells):
                    emit(r, cur_label, as_meta_kind=META_PREFIX_KIND[pm.group(1).lower()])
                else:
                    emit(r, cur_label)
                continue
            flat = collapse(raw).rstrip(":").lower()
            if flat in META_LABELS:
                if cur_label is not None:
                    emit(r, cur_label, as_meta_kind=META_LABELS[flat])
                continue
            lines = [collapse(l) for l in str(raw).split("\n") if collapse(l)]
            canon = next((c2 for ln in lines for rx, c2 in SEL_KEYWORDS if rx.match(ln)), None)
            if canon:
                cur_label = canon
                emit(r, canon)
                continue
            # any other labeled row (subject / week-note / footer / unknown)
            # closes the SEL block
            cur_label = None
    return out


# ── main assembly ────────────────────────────────────────────────────────────
def main():
    data = json.loads(EXTRACT_PATH.read_text(encoding="utf-8"))
    soft_delete_ids, unit7, standards_map, prior_units, band_names = load_prior_sql()
    sel_data = extract_sel(data["meta"]["source_xlsx"], data["meta"])

    stats = Counter()
    heading_census, combo_census = Counter(), Counter()
    no_heading_cells = []
    lessons = []          # dicts ready for SQL emission
    slug_seen = {}

    def add_lesson(gw, day, subject, title, directions, objectives, notes,
                   resources, std_uuids):
        band = band_of_week(gw)
        unit_id = prior_units.get((subject, band))
        if unit_id is None:  # spelling / ufli / sel — new unit for this band
            unit_id = slug_uuid("unit", f"g5-2627-{subject}-u{band}")
        slug = f"d-w{gw}-{subject}-{day}"
        if slug in slug_seen:  # defensive; same-subject sections merge upstream
            n = slug_seen[slug] = slug_seen[slug] + 1
            slug = f"{slug}-{n}"
        else:
            slug_seen[slug] = 1
        lessons.append({
            "slug": slug, "id": slug_uuid("lesson", slug), "unit_id": unit_id,
            "subject": subject, "week": gw, "day": day, "title": title,
            "directions": directions, "objectives": objectives, "notes": notes,
            "resources": resources, "standards": std_uuids,
        })

    for rec in data["records"]:
        gw, day, subj = rec["global_week"], rec["day_token"], rec["subject"]
        cell = rec["cell_text"]
        # span note (mapping rule 6): lesson sits on the span's first day
        covered = []
        if rec.get("merged_span"):
            a, z = rec["merged_span"]
            covered = DAY_TOKENS[DAY_TOKENS.index(a): DAY_TOKENS.index(z) + 1]
        for t in rec.get("day_span", []):
            if t not in covered:
                covered.append(t)
        span_note = span_days_note(day, covered) if covered else ""
        resources, note_rest = ([], "")
        if rec.get("label_note"):
            resources, note_rest = label_note_resources(rec["label_note"])
        base_notes = [p for p in
                      [span_note, with_kind_prefix("Note", note_rest) if note_rest else ""]
                      if p]

        if subj in ("math", "explorers"):
            notes, std_uuids, objectives = build_notes_and_standards(
                rec.get("meta"), base_notes, standards_map, stats)
            add_lesson(gw, day, subj, make_title(first_nonempty_line(cell)),
                       cell, objectives, notes, resources, std_uuids)
        elif subj == "reading":
            sections, census, combos = split_literacy_cell(cell)
            heading_census.update(census)
            combo_census.update(combos)
            if sections[0][2] is None:
                no_heading_cells.append((rec["sheet_name"], day))
                stats["literacy_cells_no_headings"] += 1
            # block-level meta + label_note attach to the ANCHOR lesson (the
            # first detected section — same one that absorbs the preamble)
            for i, (sec_subj, sec_text, heading) in enumerate(sections):
                if i == 0:
                    notes, std_uuids, objectives = build_notes_and_standards(
                        rec.get("meta"), base_notes, standards_map, stats)
                    res = resources
                else:
                    notes, std_uuids, objectives = ("", [], [])
                    if span_note:        # the whole cell spans those days
                        notes = span_note
                    res = []
                title_line = heading if heading is not None else first_nonempty_line(sec_text)
                add_lesson(gw, day, sec_subj, make_title(title_line),
                           sec_text, objectives, notes, res, std_uuids)
                stats[f"literacy_section_{sec_subj}"] += 1
        else:
            sys.exit(f"unexpected record subject {subj}")

    # SEL: one combined lesson per (week, day) — mapping rule 2. A slot can
    # exist with meta but no content (a metadata sub-row had day text while the
    # content cell was blank); only sections with real text become directions,
    # and a day with no content at all emits no lesson.
    for gw in sorted(sel_data):
        for day in sorted(sel_data[gw], key=DAY_TOKENS.index):
            slots = sel_data[gw][day]
            labels_present = [l for l in SEL_ORDER if l in slots and slots[l]["text"]]
            if not labels_present:
                stats["sel_meta_only_day_slots_skipped"] += 1
                continue
            parts, note_parts, meta_all = [], [], {}
            for label in labels_present:
                slot = slots[label]
                parts.append(f"{label}:\n{slot['text']}")
                sp = span_days_note(day, slot["covered"],
                                    label=label if len(labels_present) > 1 else None)
                if sp:
                    note_parts.append(sp)
            for label in (l for l in SEL_ORDER if l in slots):  # meta incl. text-less slots
                for k, v in slots[label]["meta"].items():
                    meta_all[k] = (meta_all[k] + "\n\n" + v) if k in meta_all else v
            notes, std_uuids, objectives = build_notes_and_standards(
                meta_all, note_parts, standards_map, stats)
            add_lesson(gw, day, "sel", " / ".join(labels_present),
                       "\n\n".join(parts), objectives, notes, [], std_uuids)

    lessons.sort(key=lambda L: (L["week"], DAY_TOKENS.index(L["day"]),
                                DISPLAY_ORDER[L["subject"]]))
    new_ids = {L["id"] for L in lessons}
    if len(new_ids) != len(lessons):
        sys.exit("duplicate lesson ids generated")
    if new_ids & set(soft_delete_ids) or new_ids & {i for i, _, _ in unit7}:
        sys.exit("new lesson id collides with a prior load id")

    # units needed for the three new subjects (bands that actually have lessons)
    new_units = sorted({(L["subject"], band_of_week(L["week"]))
                        for L in lessons if (L["subject"], band_of_week(L["week"]))
                        not in prior_units})
    subj_label = {"spelling": "Spelling", "ufli": "UFLI", "sel": "SEL"}

    # ── emit SQL ─────────────────────────────────────────────────────────────
    o = []
    w = o.append
    w("-- =============================================================================\n")
    w("-- Grade 5 — 2026-2027 DAY-LEVEL curriculum rebuild (Stage 2).\n")
    w("-- Generated by scripts/gen-grade5-daily-sql.py from\n")
    w("-- scripts/_grade5-day-extract.json (+ SEL rows re-read from the workbook).\n")
    w("--\n")
    w("-- Replaces the prior load's 185 all-on-Sunday week-summary lessons:\n")
    w(f"--   * soft-deletes the {len(soft_delete_ids)} week-1-34 summaries (explicit id list; reversible)\n")
    w(f"--   * inserts {len(lessons)} real daily lessons (idempotent upserts; re-runs heal)\n")
    w("--   * spreads the 15 week-35-37 Unit-7 Spiral Review summaries one-subject-per-day\n")
    w("--\n")
    w("-- Idempotent and safe to re-run. The whole load is ONE transaction (a\n")
    w("-- connection drop can never leave the summaries hidden with only part of the\n")
    w("-- daily rebuild inserted), and section 2's guard ABORTS everything if any\n")
    w("-- teacher fork / completion row references the lessons being soft-deleted.\n")
    w("-- DRY RUN: replace the final `commit;` with `rollback;` (see the README's\n")
    w("-- helper — do NOT wrap this file in another BEGIN, the inner commit would win).\n")
    w("-- NOTE: grade_level_id is intentionally not set on lessons — a BEFORE trigger\n")
    w("-- (planner_scale_hardening) derives it from unit_id, same as the prior load.\n")
    w("-- =============================================================================\n\n")
    w("begin;\n\n")

    w("-- ── 1. Subjects — ensure the three new planner subjects exist ──────────────\n")
    w("-- (seed.sql / seed-cloud.sql values mirrored verbatim; already present in\n")
    w("--  prod, so these are no-op insurance.)\n")
    for token, name, order in NEW_SUBJECTS:
        w("insert into subjects (id, grade_level_id, name, color, display_order, scope, default_pacing)\n")
        w(f"  values ({q(SUBJECT_ID[token])}, {q(GRADE)}, {q(name)}, {q(token)}, {order}, 'team', 'synchronized')\n")
        w("  on conflict (id) do nothing;\n")
    w("\n")

    w(f"-- ── 1b. Units for the new subjects ({len(new_units)}) ─────────────────────────────────\n")
    w("-- master_core_lesson_events.unit_id is NOT NULL, so spelling/ufli/sel lessons\n")
    w("-- need unit rows. One unit per thematic band, named like the prior load's\n")
    w("-- fallback convention (\"<band> — <Subject>\").\n")
    for subj, band in new_units:
        uid = slug_uuid("unit", f"g5-2627-{subj}-u{band}")
        a, z = next((ba, bz) for b, ba, bz in BANDS if b == band)
        name = f"{band_names[band]} — {subj_label[subj]}"
        w("insert into units (id, grade_level_id, subject_id, school_year_id, name, start_week, end_week)\n")
        w(f"  values ({q(uid)}, {q(GRADE)}, {q(SUBJECT_ID[subj])}, {q(YEAR_2627)}, {q(name)}, {a}, {z})\n")
        w("  on conflict (id) do nothing;\n")
    w("\n")

    w(f"-- ── 2. Soft-delete the {len(soft_delete_ids)} week-1-34 week-summary lessons ──────────────\n")
    w("-- Explicit id list read from scripts/_grade5-real-data.sql (slug scheme\n")
    w("-- g5-2627-<subject>-w<1..34>). Never hard-deleted; `deleted_at is null` guard\n")
    w("-- keeps re-runs from refreshing the timestamp. Reverse with:\n")
    w("--   update master_core_lesson_events set deleted_at = null where id in (<this list>);\n")
    w("--\n")
    w("-- Guard (aborts the WHOLE transaction): the 2026-06-12 pre-flight found 0\n")
    w("-- teacher-attached rows on these lessons across every dependent surface\n")
    w("-- (forks, completions, boards, lesson sections, comments, resources), but\n")
    w("-- one could land between pre-flight and apply. Locking the rows FOR UPDATE\n")
    w("-- serializes against concurrent FK writers; any reference found — or any of\n")
    w("-- the 170 rows missing — raises and rolls everything back.\n")
    w("do $$\n")
    w("declare\n")
    w("  ids uuid[] := array[\n")
    for i, lid in enumerate(soft_delete_ids):
        w(f"    {q(lid)}{',' if i < len(soft_delete_ids) - 1 else ''}\n")
    w("  ]::uuid[];\n")
    w("  n bigint;\n")
    w("begin\n")
    w("  select count(*) into n from (\n")
    w("    select id from master_core_lesson_events where id = any(ids) for update\n")
    w("  ) locked;\n")
    w("  if n <> cardinality(ids) then\n")
    w("    raise exception 'Grade 5 day-level load aborted: expected % week-summary lessons, found % — prod has drifted from scripts/_grade5-real-data.sql', cardinality(ids), n;\n")
    w("  end if;\n")
    w("  select (select count(*) from personal_core_lesson_event_copies\n")
    w("           where master_core_lesson_event_id = any(ids))\n")
    w("       + (select count(*) from completion_status\n")
    w("           where core_lesson_event_id = any(ids))\n")
    w("       + (select count(*) from boards\n")
    w("           where master_core_lesson_event_id = any(ids))\n")
    w("       + (select count(*) from lesson_sections\n")
    w("           where owner_kind = 'master' and owner_lesson_id = any(ids))\n")
    w("       + (select count(*) from comments\n")
    w("           where anchor_type = 'core_lesson_event' and anchor_id = any(ids::text[]))\n")
    w("       + (select count(*) from resources\n")
    w("           where owner_event_type = 'core_lesson_event' and owner_event_id = any(ids))\n")
    w("       into n;\n")
    w("  if n > 0 then\n")
    w("    raise exception 'Grade 5 day-level load aborted: % teacher-attached row(s) (fork/completion/board/section/comment/resource) reference the week-summary lessons being soft-deleted — resolve them first', n;\n")
    w("  end if;\n")
    w("  update master_core_lesson_events\n")
    w("     set deleted_at = now(), updated_at = now()\n")
    w("   where deleted_at is null and id = any(ids);\n")
    w("end $$;\n\n")

    w(f"-- ── 3. The {len(lessons)} day-level lessons (idempotent upserts) ──────────────────\n")
    w("-- `on conflict (id) do update` rewrites every content column AND clears\n")
    w("-- deleted_at, so re-running heals partial or previously rolled-back loads.\n")
    for L in lessons:
        std_arr = ("array[" + ", ".join(q(u) for u in L["standards"]) + "]::uuid[]"
                   if L["standards"] else "'{}'::uuid[]")
        w("insert into master_core_lesson_events\n")
        w("  (id, unit_id, subject_id, week_number, day_of_week, title, directions, learning_objectives, notes, resources, standards, display_order_within_day)\n")
        w(f"  values ({q(L['id'])}, {q(L['unit_id'])}, {q(SUBJECT_ID[L['subject']])}, "
          f"{L['week']}, {q(L['day'])}::weekday, {q(L['title'])}, {q(L['directions'])}, "
          f"{jq(L['objectives'])}, {q(L['notes'])}, {jq(L['resources'])}, {std_arr}, "
          f"{DISPLAY_ORDER[L['subject']]})\n")
        w("  on conflict (id) do update set\n")
        w("    unit_id=excluded.unit_id, subject_id=excluded.subject_id,\n")
        w("    week_number=excluded.week_number, day_of_week=excluded.day_of_week,\n")
        w("    title=excluded.title, directions=excluded.directions,\n")
        w("    learning_objectives=excluded.learning_objectives, notes=excluded.notes,\n")
        w("    resources=excluded.resources, standards=excluded.standards,\n")
        w("    display_order_within_day=excluded.display_order_within_day,\n")
        w("    deleted_at=null, updated_at=now();\n")
    w("\n")

    w("-- ── 4. Weeks 35-37 (Unit 7 Spiral Review — no workbook content) ────────────\n")
    w("-- Keep the 15 week-summary lessons; spread one subject per day. Guarded by\n")
    w("-- day_of_week='sun' so a teacher's manual move is never clobbered (and\n")
    w("-- re-runs are no-ops). math stays on Sunday — no update needed.\n")
    DAY_TARGET = {"reading": "mon", "writing": "tue", "grammar": "wed", "explorers": "thu"}
    for lid, subj, week in sorted(unit7, key=lambda t: (t[2], DISPLAY_ORDER[t[1]])):
        if subj == "math":
            w(f"-- week {week} math ({lid}) stays on 'sun'\n")
            continue
        w("update master_core_lesson_events\n")
        w(f"   set day_of_week = {q(DAY_TARGET[subj])}::weekday, updated_at = now()\n")
        w(f" where id = {q(lid)} and day_of_week = 'sun';  -- week {week} {subj}\n")
    w("\n")
    w("commit;  -- replace with `rollback;` for a dry run (README has a helper)\n\n")

    per_subj = Counter(L["subject"] for L in lessons)
    per_day = Counter(L["day"] for L in lessons)
    w("-- ── 5. Verification (run separately; kept as comments) ─────────────────────\n")
    w("-- Expected after load:\n")
    w(f"--   active day-level lessons: {len(lessons)} "
      f"({', '.join(f'{s}={per_subj[s]}' for s in DISPLAY_ORDER if per_subj[s])})\n")
    w(f"--   per day: {', '.join(f'{d}={per_day[d]}' for d in DAY_TOKENS if per_day[d])}\n")
    w(f"--   plus 15 active week-35-37 summaries = {len(lessons) + 15} active total; "
      f"{len(soft_delete_ids)} soft-deleted.\n")
    w("--\n")
    w("-- select s.color as subject, m.day_of_week, count(*)\n")
    w("--   from master_core_lesson_events m join subjects s on s.id = m.subject_id\n")
    w("--   join units u on u.id = m.unit_id\n")
    w("--  where u.school_year_id = '00000000-0000-0000-0000-0000000000c2'\n")
    w("--    and m.deleted_at is null\n")
    w("--  group by 1, 2 order by 1, 2;\n")
    w("--\n")
    w("-- select count(*) as active_total\n")
    w("--   from master_core_lesson_events m join units u on u.id = m.unit_id\n")
    w("--  where u.school_year_id = '00000000-0000-0000-0000-0000000000c2'\n")
    w(f"--    and m.deleted_at is null;                       -- expect {len(lessons) + 15}\n")
    w("--\n")
    w("-- select count(*) as soft_deleted\n")
    w("--   from master_core_lesson_events m join units u on u.id = m.unit_id\n")
    w("--  where u.school_year_id = '00000000-0000-0000-0000-0000000000c2'\n")
    w(f"--    and m.deleted_at is not null;                   -- expect {len(soft_delete_ids)}\n")
    w("--\n")
    w("-- select day_of_week, count(*) from master_core_lesson_events\n")
    w("--  where week_number between 35 and 37 and deleted_at is null\n")
    w("--  group by 1 order by 1;                             -- expect 3 per weekday\n")

    OUT_PATH.write_text("".join(o), encoding="utf-8")

    # ── run report ───────────────────────────────────────────────────────────
    print(f"wrote {OUT_PATH}")
    print(f"lessons={len(lessons)} soft_delete={len(soft_delete_ids)} "
          f"unit7_updates={sum(1 for _, s, _ in unit7 if s != 'math')} "
          f"new_units={len(new_units)} subjects_upserted={len(NEW_SUBJECTS)}")
    print("\nper subject:", {s: per_subj[s] for s in DISPLAY_ORDER if per_subj[s]})
    print("per day:   ", {d: per_day[d] for d in DAY_TOKENS if per_day[d]})
    per_week = Counter(L["week"] for L in lessons)
    counts = sorted(per_week.values())
    print(f"per week:   min={counts[0]} avg={sum(counts)/len(counts):.1f} "
          f"max={counts[-1]} weeks={len(counts)}")
    print("\nliteracy heading census:", dict(heading_census.most_common()))
    print("combined headings:      ", dict(combo_census.most_common()))
    print(f"literacy cells with no headings: {stats['literacy_cells_no_headings']}",
          (no_heading_cells if no_heading_cells else ""))
    print("standards: matched piece->code hits =", stats["std_matched"],
          "| valid-shaped but not in prior map =", stats["std_unmatched_code"],
          "| lessons keeping full std text in notes =", stats["std_text_kept_in_notes"])
    empty_t = [L["slug"] for L in lessons if not L["title"].strip()]
    empty_d = [L["slug"] for L in lessons if not L["directions"].strip()]
    print("empty titles:", empty_t or "none", "| empty directions:", empty_d or "none")
    print("sel meta-only day-slots skipped:", stats["sel_meta_only_day_slots_skipped"])
    big = sorted(lessons, key=lambda L: len(L["directions"]), reverse=True)[:2]
    for L in big:
        print(f"largest: {L['slug']} ({len(L['directions'])} chars) title={L['title']!r}")


if __name__ == "__main__":
    main()

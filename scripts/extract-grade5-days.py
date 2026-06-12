#!/usr/bin/env python3
# =============================================================================
# extract-grade5-days.py — Stage-1 extractor for the Grade 5 per-day rebuild.
#
# Parses the teacher workbook "Grade 5 Weekly Plans 2026-2027.xlsx" (one sheet
# per unit-week, column A = class/row label, columns B-F = Sunday..Thursday)
# into a clean per-day lesson inventory JSON at scripts/_grade5-day-extract.json.
#
# Conventions mirror scripts/gen-grade5-real-data.py (the canonical reference):
#   - global weeks 1..37 across units (workbook covers 1..34; the generator's
#     "Unit 7: Natural Disasters and Spiral Review" weeks 35-37 have no sheets)
#   - app subject tokens: math, reading, writing, grammar, explorers
#
# IMPORTANT mapping caveat (teacher approval required before anything ships):
# this workbook has NO separate Reading / Writing / Grammar rows. A single
# "Literacy" row carries the whole ELA block; Writing/Grammar/Spelling/Phonics
# live INSIDE its day cells as inline headings. The proposed (draft) mapping
# emits Literacy rows as subject "reading" with a `sections_detected` hint so
# a later stage can split — see PROPOSED_MAPPING below and the run report.
#
# Deterministic, offline, stdlib + openpyxl only.
#   pip install --user openpyxl
# =============================================================================

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install --user openpyxl")

# ── configuration (DATA_DIR convention follows gen-grade5-real-data.py) ──────
DATA_DIR = r"C:\Users\losey\OneDrive\Documents\Claude\Projects\Grade 5 Curriculum"
XLSX_PATH = rf"{DATA_DIR}\Grade 5 Weekly Plans 2026-2027.xlsx"
OUT_PATH = str(Path(__file__).resolve().parent / "_grade5-day-extract.json")

# Canonical day order for this school's Sun-Thu week (sample data, not a
# product assumption — the app's school week is configurable).
DAY_TOKENS = ["sun", "mon", "tue", "wed", "thu", "fri", "sat"]
DAY_NAMES = {
    "sunday": "sun", "monday": "mon", "tuesday": "tue",
    "wednesday": "wed", "thursday": "thu", "friday": "fri", "saturday": "sat",
}
SUBJECT_ORDER = {"math": 0, "reading": 1, "writing": 2, "grammar": 3, "explorers": 4}

# ── row-label classification ─────────────────────────────────────────────────
# Subject block starts: the first line of the label cell that matches one of
# these keywords decides the block; any remaining label text is preserved as
# `label_note` (teachers sometimes type week notes / padlet links into the
# label cell itself).
SUBJECT_KEYWORDS = [
    # (regex on a single collapsed label line, canonical label, app subject)
    (re.compile(r"^literacy\b", re.I), "Literacy", "reading"),
    (re.compile(r"^math\b", re.I), "Math", "math"),
    (re.compile(r"^explore\b", re.I), "Explore", "explorers"),
    (re.compile(r"^ipc\b", re.I), "Explore", "explorers"),  # IPC == Explore
]
# Non-lesson blocks (SEL / routines). Not one of the 5 active app subjects →
# captured in the "unmapped" list, never silently dropped.
UNMAPPED_KEYWORDS = [
    (re.compile(r"^morning meeting\b", re.I), "Morning Meeting"),
    (re.compile(r"^closing circle\b", re.I), "Closing Circle (SELG)"),
    (re.compile(r"^se learning\b", re.I), "SE Learning PL Goal"),
]
# Metadata sub-rows: attach to the preceding block, keyed by kind.
META_LABELS = {
    "learning goals objectives": "objectives",
    "learning goals/objectives": "objectives",
    "learning objectives": "objectives",
    "objective": "objectives",
    "objectives": "objectives",
    "assessment": "assessment",
    "assessments": "assessment",
    "standard": "standards",
    "standards": "standards",
    "differentiation": "differentiation",
    "personal learning goal/islamic values": "plg_islamic_values",
    "ipc personal learning goal/islamic values": "plg_islamic_values",
}
# Week-level rows (not lessons, not block metadata).
WEEK_NOTE_LABELS = {"this weeks updates/notes", "this week's updates/notes"}
FOOTER_LABELS = {"notes"}

# Inline ELA section markers inside Literacy cells (advisory hint only — the
# full cell text always travels with the record).
SECTION_MARKERS = {
    "guided_reading": re.compile(r"\bguided\s*read\w*\s*[:/\-–]", re.I),
    "writing": re.compile(r"\bwriting\b\s*[:/\-–]", re.I),
    "grammar": re.compile(r"\bgrammar\b\s*[:/\-–]", re.I),
    "spelling": re.compile(r"\bspelling\b\s*[:/\-–]?", re.I),
    "phonics": re.compile(r"\bphonics\b\s*[:/\-–]?|\bdecod\w*", re.I),
    "word_study": re.compile(r"\bword\s*(study|work)\b", re.I),
}
META_CELL_PREFIX = re.compile(
    r"^\s*(assessment|assessments|obj|objective|differentiation|std|standard|plg)\b[\s:.-]",
    re.I,
)
META_PREFIX_KIND = {
    "assessment": "assessment", "assessments": "assessment",
    "obj": "objectives", "objective": "objectives",
    "differentiation": "differentiation",
    "std": "standards", "standard": "standards",
    "plg": "plg_islamic_values",
}

# Proposed row-label → app-subject mapping, embedded for the approval review.
PROPOSED_MAPPING = {
    "subjects": {
        "Math": "math",
        "Literacy": "reading  (JUDGMENT CALL: single Literacy row carries the whole "
                    "ELA block — Guided Reading + Writing + Grammar + Spelling/Phonics "
                    "inline. No dedicated writing/grammar rows exist in this workbook; "
                    "see sections_detected on each record.)",
        "Explore / IPC": "explorers",
    },
    "unmapped_pending_approval": {
        "Morning Meeting / Morning Meeting/SEL": "SEL/routine — 'sel' is not one of "
            "the 5 active planner subjects for this grade",
        "Closing Circle (SELG) / Closing Circle Reminders": "SEL/routine",
        "SE Learning PL Goal / SE Learning": "weekly SEL lesson row",
    },
    "metadata_not_lessons": [
        "Learning Goals Objectives", "Learning Objectives", "Objective(s)",
        "Assessment(s)", "Standards / Standard", "Differentiation",
        "Personal Learning Goal/Islamic Values (incl. IPC variant)",
        "This weeks updates/notes:", "Notes:",
    ],
}


# ── helpers ──────────────────────────────────────────────────────────────────
def collapse(s):
    return re.sub(r"\s+", " ", str(s)).strip()


def clean_text(s):
    """Trim + normalize whitespace, preserve bullets/lines."""
    if s is None:
        return ""
    t = str(s).replace("\r\n", "\n").replace("\r", "\n")
    lines = [ln.rstrip() for ln in t.split("\n")]
    t = "\n".join(lines)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def classify_label(raw):
    """→ ("subject", canonical, token, note) | ("unmapped", canonical, note)
       | ("meta", kind) | ("weeknote",) | ("footer",) | ("unknown", collapsed)"""
    flat = collapse(raw).rstrip(":").lower()
    if not flat:
        return ("blank",)
    if flat in META_LABELS:
        return ("meta", META_LABELS[flat])
    if flat in WEEK_NOTE_LABELS:
        return ("weeknote",)
    if flat in FOOTER_LABELS:
        return ("footer",)
    # line-scan: first line matching a subject / unmapped keyword wins; the
    # rest of the label cell becomes the note.
    lines = [collapse(ln) for ln in str(raw).split("\n")]
    lines = [ln for ln in lines if ln]
    for i, ln in enumerate(lines):
        low = ln.rstrip(":").lower()
        if low in META_LABELS:  # e.g. "IPC Personal Learning Goal/Islamic Values"
            return ("meta", META_LABELS[low])
        for rx, canon, token in SUBJECT_KEYWORDS:
            if rx.match(ln):
                note = collapse(" ".join(lines[:i] + [rx.sub("", ln, count=1)] + lines[i + 1:]))
                return ("subject", canon, token, note)
        for rx, canon in UNMAPPED_KEYWORDS:
            if rx.match(ln):
                note = collapse(" ".join(lines[:i] + [rx.sub("", ln, count=1)] + lines[i + 1:]))
                return ("unmapped", canon, note)
    return ("unknown", collapse(raw))


def find_header_row(ws):
    for row in ws.iter_rows(min_row=1, max_row=25, max_col=1):
        v = row[0].value
        if v and collapse(v).lower() == "class":
            return row[0].row
    return None


def parse_day_columns(ws, hdr):
    """[(col, [day tokens covered])] from the header row; 'SundayMonday' → 2 tokens."""
    out = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if v is None or not collapse(v):
            continue
        name = collapse(v).lower()
        toks = [tok for dn, tok in DAY_NAMES.items() if dn in name]
        toks.sort(key=DAY_TOKENS.index)
        if toks:
            out.append((c, toks))
    return out


def build_merge_maps(ws):
    """anchor lookup {(r,c): (ar,ac)} + list of merge ranges."""
    amap = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                amap[(r, c)] = (mr.min_row, mr.min_col)
    return amap, list(ws.merged_cells.ranges)


# ── main extraction ──────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)  # not read_only: need merges
    week_re = re.compile(r"^U(\d+) W(\d+)$")
    week_sheets = []
    for n in wb.sheetnames:
        m = week_re.match(n)
        if m:
            week_sheets.append((int(m.group(1)), int(m.group(2)), n))
    week_sheets.sort()

    # unit names from the "U<N> Summary" sheets ("Unit N: <name>")
    unit_names = {}
    for n in wb.sheetnames:
        m = re.match(r"^U(\d+) Summary$", n)
        if m:
            title = collapse(wb[n].cell(row=1, column=1).value or "")
            unit_names[int(m.group(1))] = re.sub(r"^Unit\s*\d+\s*:\s*", "", title)

    # global week offsets: cumulative unit week counts in unit order (mirrors
    # gen-grade5-real-data.py, whose CSV-derived units 1-6 have identical counts)
    weeks_per_unit = Counter(u for u, _, _ in week_sheets)
    offsets, acc = {}, 0
    for u in sorted(weeks_per_unit):
        offsets[u] = acc
        acc += weeks_per_unit[u]

    records = []
    unmapped = defaultdict(lambda: {"count_rows": 0, "sheets": [], "sample_text": "",
                                    "label_variants": set(), "meta_kinds": set()})
    unknown_rows = []
    anomalies = []
    label_stats = Counter()
    sheet_meta = {}

    for u, w, sn in week_sheets:
        ws = wb[sn]
        gw = offsets[u] + w
        hdr = find_header_row(ws)
        if hdr is None:
            anomalies.append({"sheet": sn, "issue": "no 'Class' header row found — sheet skipped"})
            continue
        day_cols = parse_day_columns(ws, hdr)
        if [t for _, toks in day_cols for t in toks] != ["sun", "mon", "tue", "wed", "thu"]:
            anomalies.append({"sheet": sn, "issue": f"non-standard day header: "
                              f"{[(c, toks) for c, toks in day_cols]} (header row {hdr})"})
        amap, merges = build_merge_maps(ws)

        def cell_val(r, c):
            """value honoring merges; (value, anchor_row, anchor_col, merge or None)"""
            if (r, c) in amap:
                ar, ac = amap[(r, c)]
                mr = next(m for m in merges
                          if m.min_row == ar and m.min_col == ac
                          and m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col)
                return ws.cell(row=ar, column=ac).value, ar, ac, mr
            return ws.cell(row=r, column=c).value, r, c, None

        # sheet-level meta: title, dates, "Weeks: N of M", preamble
        title = collapse(ws.cell(row=1, column=1).value or "")
        dates_text, weeks_label = "", ""
        for r in range(2, hdr):
            a = collapse(ws.cell(row=r, column=1).value or "")
            if a.lower().startswith("dates:"):
                rest = a[len("dates:"):].strip()
                dates_text = rest or collapse(ws.cell(row=r, column=2).value or "")
            for c in range(1, ws.max_column):
                if collapse(ws.cell(row=r, column=c).value or "").lower() == "weeks:":
                    weeks_label = collapse(ws.cell(row=r, column=c + 1).value or "")
        m = re.match(r"^(\d+)\s*of\s*(\d+)$", weeks_label)
        if m and (int(m.group(1)) != w or int(m.group(2)) != weeks_per_unit[u]):
            anomalies.append({"sheet": sn, "issue": f"'Weeks: {weeks_label}' disagrees with "
                              f"sheet position W{w} of {weeks_per_unit[u]}"})

        week_notes, footer_note = {}, ""

        # block-state walk over rows below the header
        cur = None  # {"kind": "subject"|"unmapped", "token", "canon", "note"}
        # per (subject, day_token): accumulated content segments + meta
        content = defaultdict(list)   # (token, day) -> [text]
        meta = defaultdict(dict)      # (token, day) -> {kind: text}
        meta_week_scope = defaultdict(set)  # token -> {kinds merged across the week}
        block_rows = defaultdict(list)      # token -> [sheet rows]
        spans = {}                    # (token, day) -> [first, last] merged day span
        day_spans = {}                # (token, day) -> combined-header tokens

        def emit_day_cells(row, sink_token, as_meta_kind=None):
            """walk day columns of `row`, honoring horizontal merges."""
            seen_anchors = set()
            for c, toks in day_cols:
                v, ar, ac, mr = cell_val(row, c)
                if v is None or not clean_text(v):
                    continue
                if ar != row:  # vertical merge bleeding in from another row
                    if (ar, ac) not in seen_anchors:
                        anomalies.append({"sheet": sn, "issue": f"vertical merge {mr} bleeds "
                                          f"into row {row} (col {c}) — counted at row {ar} only"})
                        seen_anchors.add((ar, ac))
                    continue
                if (ar, ac) in seen_anchors:
                    continue  # horizontal merge already captured at its first column
                seen_anchors.add((ar, ac))
                covered = [t for cc, tt in day_cols for t in tt
                           if mr is not None and mr.min_col <= cc <= mr.max_col] \
                          if mr is not None else list(toks)
                first_tok = toks[0]
                txt = clean_text(v)
                if as_meta_kind:
                    tgt_days = covered if mr is not None and len(covered) > len(toks) else toks
                    if mr is not None and len(covered) > len(toks):
                        meta_week_scope[sink_token].add(as_meta_kind)
                    for t in tgt_days:
                        prev = meta[(sink_token, t)].get(as_meta_kind)
                        meta[(sink_token, t)][as_meta_kind] = (prev + "\n\n" + txt) if prev else txt
                else:
                    content[(sink_token, first_tok)].append(txt)
                    if mr is not None and len(covered) > len(toks):
                        spans[(sink_token, first_tok)] = [covered[0], covered[-1]]
                    if len(toks) > 1:
                        day_spans[(sink_token, first_tok)] = list(toks)

        for r in range(hdr + 1, ws.max_row + 1):
            raw_label, lar, lac, lmr = cell_val(r, 1)
            if lar != r:
                # label cell merged down from a previous row: this row continues
                # the current block — do not re-classify or re-count the label.
                raw_label = None
            cls = classify_label(raw_label) if raw_label is not None else ("blank",)
            if cls[0] == "blank":
                # unlabeled row — content may continue the current block
                cells = []
                for c, _ in day_cols:
                    v, ar, _, _ = cell_val(r, c)
                    if ar == r and v is not None and clean_text(v):
                        cells.append(clean_text(v))
                if not cells:
                    continue
                if cur is None:
                    anomalies.append({"sheet": sn, "issue": f"row {r}: content with no label "
                                      "and no open block — skipped", "sample": cells[0][:120]})
                    continue
                pm = META_CELL_PREFIX.match(cells[0])
                if pm and all(META_CELL_PREFIX.match(x) for x in cells):
                    kind = META_PREFIX_KIND[pm.group(1).lower()]
                    emit_day_cells(r, cur["token"], as_meta_kind=kind)
                    if cur["kind"] == "unmapped":
                        unmapped[cur["canon"]]["meta_kinds"].add(kind)
                else:
                    emit_day_cells(r, cur["token"])  # content continuation
                    if cur["kind"] == "subject":
                        block_rows[cur["token"]].append(r)
                        anomalies.append({"sheet": sn, "issue": f"row {r}: unlabeled content row "
                                          f"appended to '{cur['canon']}' block"})
                continue

            label_stats[collapse(raw_label)] += 1
            if cls[0] == "weeknote":
                for c, toks in day_cols:
                    v, ar, _, _ = cell_val(r, c)
                    if v is not None and clean_text(v):
                        week_notes[toks[0]] = clean_text(v)
                cur = None
                continue
            if cls[0] == "footer":
                vals = [clean_text(ws.cell(row=r, column=c).value)
                        for c in range(2, ws.max_column + 1)
                        if ws.cell(row=r, column=c).value]
                footer_note = "\n".join(v for v in vals if v)
                cur = None
                continue
            if cls[0] == "meta":
                if cur is None:
                    anomalies.append({"sheet": sn, "issue": f"row {r}: metadata row "
                                      f"'{collapse(raw_label)}' with no open block — skipped"})
                    continue
                emit_day_cells(r, cur["token"], as_meta_kind=cls[1])
                if cur["kind"] == "unmapped":
                    unmapped[cur["canon"]]["meta_kinds"].add(cls[1])
                continue
            if cls[0] == "subject":
                _, canon, token, note = cls
                cur = {"kind": "subject", "token": token, "canon": canon, "note": note}
                block_rows[token].append(r)
                if note:
                    meta[(token, "_label_note")]["note"] = note
                emit_day_cells(r, token)
                continue
            if cls[0] == "unmapped":
                _, canon, note = cls
                bucket = unmapped[canon]
                bucket["count_rows"] += 1
                bucket["sheets"].append(sn)
                bucket["label_variants"].add(collapse(raw_label))
                if not bucket["sample_text"]:
                    for c, _ in day_cols:
                        v, ar, _, _ = cell_val(r, c)
                        if ar == r and v is not None and clean_text(v):
                            bucket["sample_text"] = clean_text(v)[:600]
                            break
                cur = {"kind": "unmapped", "token": f"__unmapped__{canon}", "canon": canon}
                continue
            # unknown label
            unknown_rows.append({"sheet": sn, "row": r, "row_label": cls[1][:200],
                                 "sample_text": next((clean_text(cell_val(r, c)[0])[:300]
                                                      for c, _ in day_cols
                                                      if cell_val(r, c)[0]), "")})
            cur = None

        # materialize records for this sheet
        for (token, day), segs in sorted(content.items()):
            if token.startswith("__unmapped__") or day == "_label_note":
                continue
            txt = "\n\n".join(segs)
            if not txt:
                continue
            canon = {"reading": "Literacy", "math": "Math", "explorers": "Explore"}[token]
            rec = {
                "unit_index": u,
                "unit_name": unit_names.get(u, ""),
                "global_week": gw,
                "sheet_name": sn,
                "row_label": canon,
                "subject": token,
                "day_token": day,
                "day_index": DAY_TOKENS.index(day),
                "cell_text": txt,
            }
            note = meta.get((token, "_label_note"), {}).get("note")
            if note:
                rec["label_note"] = note
            if (token, day) in spans:
                rec["merged_span"] = spans[(token, day)]
            if (token, day) in day_spans:
                rec["day_span"] = day_spans[(token, day)]
            if block_rows.get(token):
                rec["source_rows"] = block_rows[token]
            if token == "reading":
                hits = sorted(k for k, rx in SECTION_MARKERS.items() if rx.search(txt))
                rec["sections_detected"] = hits
            m_ = {k: v for k, v in meta.get((token, day), {}).items()}
            if m_:
                rec["meta"] = m_
                wk = sorted(meta_week_scope.get(token, ()))
                if wk:
                    rec["meta_week_scoped"] = wk
            records.append(rec)

        sheet_meta[sn] = {
            "unit_index": u, "week_in_unit": w, "global_week": gw, "title": title,
            "dates_text": dates_text, "weeks_label": weeks_label,
            "week_notes": week_notes, **({"footer_note": footer_note} if footer_note else {}),
        }

    records.sort(key=lambda r: (r["global_week"], SUBJECT_ORDER[r["subject"]], r["day_index"]))

    # ── coverage stats ───────────────────────────────────────────────────────
    slots_per_sheet = {sn: len(parse_day_columns(wb[sn], find_header_row(wb[sn])))
                       for _, _, sn in week_sheets}
    coverage = {}
    by_subj_day = defaultdict(set)
    days_covered = defaultdict(set)
    for r in records:
        by_subj_day[r["subject"]].add((r["sheet_name"], r["day_token"]))
        cov = r.get("merged_span")
        toks = (DAY_TOKENS[DAY_TOKENS.index(cov[0]): DAY_TOKENS.index(cov[1]) + 1]
                if cov else [r["day_token"]])
        toks += r.get("day_span", [])
        for t in toks:
            days_covered[r["subject"]].add((r["sheet_name"], t))
    for token in ["math", "reading", "explorers"]:
        total = sum(slots_per_sheet.values())
        nonempty = len(by_subj_day[token])
        coverage[token] = {
            "records": sum(1 for r in records if r["subject"] == token),
            "day_slots_total": total,
            "day_slots_nonempty": nonempty,
            "day_slots_empty": total - nonempty,
            "days_covered_incl_merges": len(days_covered[token]),
        }
    for token in ["writing", "grammar"]:
        coverage[token] = {"records": 0, "day_slots_total": 0, "day_slots_nonempty": 0,
                           "day_slots_empty": 0,
                           "note": "no dedicated row in workbook — content embedded in "
                                   "the Literacy row (see sections_detected)"}

    out = {
        "meta": {
            "source_xlsx": XLSX_PATH,
            "generated_by": "scripts/extract-grade5-days.py",
            "week_sheet_count": len(week_sheets),
            "units": {str(u): {"name": unit_names.get(u, ""), "weeks": weeks_per_unit[u],
                               "global_weeks": [offsets[u] + 1, offsets[u] + weeks_per_unit[u]]}
                      for u in sorted(weeks_per_unit)},
            "global_weeks_covered": [1, acc],
            "generator_reconciliation": (
                "gen-grade5-real-data.py builds 37 global weeks from the curriculum-site "
                "CSV: units 1-6 (5,7,6,4,4,8 weeks = 1..34, identical to this workbook) "
                "plus 'Unit 7: Natural Disasters and Spiral Review' weeks 35-37, which "
                "has NO sheets in this workbook."),
            "proposed_mapping": PROPOSED_MAPPING,
            "day_tokens": DAY_TOKENS[:5],
            "field_semantics": {
                "unit_index": "1-based, matches the workbook's U<N> sheets and the "
                              "generator's g5-2627-<subject>-u<N> slugs",
                "cell_text": "cleaned cell value; multi-row subject blocks are joined "
                             "with a blank line in row order (see source_rows)",
                "label_note": "extra text the teacher typed into the row-label cell "
                              "(column A) beyond the subject keyword",
                "merged_span": "[first_day, last_day] — at least one source cell was "
                               "merged across these day columns; the text applies to "
                               "the whole span, and no separate records exist for the "
                               "covered days unless another source row supplied them",
                "day_span": "header-level combined day column (U6 W4 'SundayMonday')",
                "sections_detected": "Literacy only: inline ELA section headings found "
                                     "in cell_text (advisory; full text is authoritative)",
                "meta": "same-day text from the block's metadata sub-rows "
                        "(standards / objectives / assessment / differentiation / "
                        "plg_islamic_values)",
                "meta_week_scoped": "meta kinds whose source cell was merged across "
                                    "the week (replicated onto every day's record)",
            },
            "coverage": coverage,
        },
        "sheets": sheet_meta,
        "records": records,
        "unmapped": [
            {"row_label": canon,
             "count_rows": b["count_rows"],
             "sheets": sorted(set(b["sheets"]),
                              key=lambda s: (int(re.match(r"U(\d+) W(\d+)", s).group(1)),
                                             int(re.match(r"U(\d+) W(\d+)", s).group(2)))),
             "label_variants": sorted(b["label_variants"]),
             "meta_kinds_attached": sorted(b["meta_kinds"]),
             "sample_text": b["sample_text"]}
            for canon, b in sorted(unmapped.items())
        ],
        "unknown_rows": unknown_rows,
        "label_stats": dict(label_stats.most_common()),
        "anomalies": anomalies,
    }
    Path(OUT_PATH).write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"sheets={len(week_sheets)} records={len(records)} "
          f"unmapped_labels={len(unmapped)} unknown_rows={len(unknown_rows)} "
          f"anomalies={len(anomalies)}")
    for token in ["math", "reading", "writing", "grammar", "explorers"]:
        print(f"  {token:10s} {coverage[token]}")
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
